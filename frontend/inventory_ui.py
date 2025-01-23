import sys
import os
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QVBoxLayout, QHBoxLayout, QTableWidget,
    QTableWidgetItem, QPushButton, QLineEdit, QLabel, QWidget, QMessageBox, 
    QHeaderView, QFrame, QGroupBox, QFormLayout, QSpinBox, QDoubleSpinBox,
    QTextEdit, QComboBox, QDateEdit, QFileDialog, QTabWidget, QDialog, QDialogButtonBox
)
from PyQt5.QtGui import QPalette, QColor
from PyQt5.QtCore import Qt, QDate
from openpyxl import Workbook
import sqlite3
from utils.barcode_utils import generar_codigo_desde_db, imprimir_codigo_barras


class InventoryApp(QWidget):
    def __init__(self, user_data=None):
        super().__init__()
        self.user_data = user_data
        self.setup_ui()
        self.cargar_inventario()

    def setup_ui(self):
        layout = QVBoxLayout(self)
        
        # Crear el widget de pestañas
        self.tabs = QTabWidget()
        
        # Pestaña de Inventario Activo
        self.tab_inventario = QWidget()
        self.setup_inventario_tab()
        self.tabs.addTab(self.tab_inventario, "Inventario Activo")
        
        # Pestaña de Historial
        self.tab_historial = QWidget()
        self.setup_historial_tab()
        self.tabs.addTab(self.tab_historial, "Historial de Inventario")
        
        layout.addWidget(self.tabs)

    def setup_inventario_tab(self):
        layout = QVBoxLayout(self.tab_inventario)
        
        # Botones superiores
        button_layout = QHBoxLayout()
        
        btn_agregar = QPushButton("Agregar Artículo")
        btn_agregar.clicked.connect(self.agregar_articulo)
        btn_editar = QPushButton("Editar Artículo")
        btn_editar.clicked.connect(self.mostrar_editar_articulo)
        
        button_layout.addWidget(btn_agregar)
        button_layout.addWidget(btn_editar)
        button_layout.addStretch()
        
        layout.addLayout(button_layout)
        
        # Campos de entrada
        input_layout = QHBoxLayout()
        self.input_widgets = {}
        
        for label in ["Nombre:", "Descripción:", "Cantidad:", 
                     "Valor Real:", "Valor Venta:", "Categoría:"]:
            input_group = QVBoxLayout()
            input_group.addWidget(QLabel(label))
            input_widget = QLineEdit()
            input_group.addWidget(input_widget)
            self.input_widgets[label] = input_widget
            input_layout.addLayout(input_group)
        
        layout.addLayout(input_layout)
        
        # Tabla de inventario
        self.tabla_inventario = QTableWidget()
        self.tabla_inventario.setColumnCount(10)
        self.tabla_inventario.setHorizontalHeaderLabels([
            "ID", "Nombre", "Descripción", "Cantidad",
            "Precio Compra", "Precio Venta", "Categoría",
            "Código Barras", "Fecha Creación", "Creado por"
        ])
        
        # Permitir selección de fila completa
        self.tabla_inventario.setSelectionBehavior(QTableWidget.SelectRows)
        self.tabla_inventario.setSelectionMode(QTableWidget.SingleSelection)
        
        header = self.tabla_inventario.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.ResizeToContents)
        layout.addWidget(self.tabla_inventario)
        
        self.cargar_inventario()

    def setup_historial_tab(self):
        layout = QVBoxLayout(self.tab_historial)
        
        # Filtros superiores
        filter_layout = QHBoxLayout()
        
        # Filtro por fecha
        self.fecha_desde = QDateEdit()
        self.fecha_desde.setCalendarPopup(True)
        self.fecha_desde.setDate(QDate.currentDate().addMonths(-1))
        
        self.fecha_hasta = QDateEdit()
        self.fecha_hasta.setCalendarPopup(True)
        self.fecha_hasta.setDate(QDate.currentDate())
        
        # Búsqueda
        self.buscar_articulo = QLineEdit()
        self.buscar_articulo.setPlaceholderText("Buscar por nombre o descripción...")
        self.buscar_articulo.textChanged.connect(self.filtrar_historial)
        
        # Filtro por categoría
        self.categoria_combo = QComboBox()
        self.categoria_combo.addItem("Todas las categorías")
        
        filter_layout.addWidget(QLabel("Desde:"))
        filter_layout.addWidget(self.fecha_desde)
        filter_layout.addWidget(QLabel("Hasta:"))
        filter_layout.addWidget(self.fecha_hasta)
        filter_layout.addWidget(QLabel("Categoría:"))
        filter_layout.addWidget(self.categoria_combo)
        filter_layout.addWidget(self.buscar_articulo)
        
        btn_buscar = QPushButton("Buscar")
        btn_buscar.clicked.connect(self.cargar_historial)
        filter_layout.addWidget(btn_buscar)
        
        layout.addLayout(filter_layout)
        
        # Tabla de historial
        self.historial_table = QTableWidget()
        self.historial_table.setColumnCount(11)
        self.historial_table.setHorizontalHeaderLabels([
            "ID", "Nombre", "Descripción", "Stock Inicial",
            "Precio Compra", "Precio Venta", "Categoría",
            "Código Barras", "Fecha Creación", "Creado por",
            "Estado Actual"
        ])
        
        header = self.historial_table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.ResizeToContents)
        layout.addWidget(self.historial_table)
        
        # Botón exportar
        btn_exportar = QPushButton("Exportar a Excel")
        btn_exportar.clicked.connect(self.exportar_historial)
        layout.addWidget(btn_exportar)
        
        # Cargar datos iniciales
        self.cargar_categorias()
        self.cargar_historial()

    def cargar_categorias(self):
        try:
            conn = sqlite3.connect("db/store.db")
            cursor = conn.cursor()
            
            cursor.execute("""
                SELECT DISTINCT categoria 
                FROM inventario 
                WHERE categoria IS NOT NULL 
                AND categoria != ''
            """)
            
            categorias = cursor.fetchall()
            self.categoria_combo.clear()
            self.categoria_combo.addItem("Todas las categorías")
            
            for categoria in categorias:
                self.categoria_combo.addItem(categoria[0])
            
            conn.close()
            
        except sqlite3.Error as e:
            QMessageBox.critical(self, "Error", f"Error al cargar categorías: {str(e)}")

    def cargar_historial(self):
        try:
            conn = sqlite3.connect("db/store.db")
            cursor = conn.cursor()
            
            query = """
                SELECT 
                    i.id, i.nombre, i.descripcion, i.cantidad,
                    i.precio_compra, i.precio_venta, i.categoria,
                    i.codigo_barras, i.fecha_creacion,
                    COALESCE(u.telefono, 'Sistema') as creador,
                    CASE 
                        WHEN i.cantidad > 0 THEN 'En Stock'
                        ELSE 'Sin Stock'
                    END as estado
                FROM inventario i
                LEFT JOIN usuarios u ON i.usuario_creacion = u.id
                WHERE i.fecha_creacion BETWEEN ? AND ?
            """
            
            params = [
                self.fecha_desde.date().toString("yyyy-MM-dd"),
                self.fecha_hasta.date().toString("yyyy-MM-dd")
            ]
            
            # Filtrar por categoría si se seleccionó una
            categoria = self.categoria_combo.currentText()
            if categoria != "Todas las categorías":
                query += " AND i.categoria = ?"
                params.append(categoria)
            
            cursor.execute(query + " ORDER BY i.fecha_creacion DESC", params)
            items = cursor.fetchall()
            
            self.historial_table.setRowCount(len(items))
            
            for row, item in enumerate(items):
                for col, value in enumerate(item):
                    if col in [4, 5]:  # Valores monetarios
                        text = f"${value:.2f}"
                    elif col == 8:  # Fecha
                        text = value.split()[0] if value else ''
                    else:
                        text = str(value)
                    table_item = QTableWidgetItem(text)
                    table_item.setTextAlignment(Qt.AlignCenter)
                    self.historial_table.setItem(row, col, table_item)
            
            conn.close()
            
        except sqlite3.Error as e:
            QMessageBox.critical(self, "Error", f"Error al cargar historial: {str(e)}")

    def filtrar_historial(self):
        texto = self.buscar_articulo.text().lower()
        for row in range(self.historial_table.rowCount()):
            mostrar = False
            # Buscar en nombre y descripción
            for col in [1, 2]:  # Columnas de nombre y descripción
                item = self.historial_table.item(row, col)
                if item and texto in item.text().lower():
                    mostrar = True
                    break
            self.historial_table.setRowHidden(row, not mostrar)

    def exportar_historial(self):
        try:
            filename, _ = QFileDialog.getSaveFileName(
                self, 
                "Guardar Excel", 
                "", 
                "Excel Files (*.xlsx)"
            )
            
            if filename:
                if not filename.endswith('.xlsx'):
                    filename += '.xlsx'
                
                wb = Workbook()
                ws = wb.active
                ws.title = "Historial de Inventario"
                
                # Escribir encabezados
                headers = [
                    "ID", "Nombre", "Descripción", "Stock Inicial",
                    "Precio Compra", "Precio Venta", "Categoría",
                    "Código Barras", "Fecha Creación", "Creado por",
                    "Estado Actual"
                ]
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)
                
                # Escribir datos
                for row in range(self.historial_table.rowCount()):
                    if not self.historial_table.isRowHidden(row):
                        for col in range(self.historial_table.columnCount()):
                            item = self.historial_table.item(row, col)
                            ws.cell(row=row+2, column=col+1, value=item.text())
                
                wb.save(filename)
                QMessageBox.information(
                    self, 
                    "Éxito", 
                    "Historial exportado correctamente"
                )
                
        except Exception as e:
            QMessageBox.critical(
                self, 
                "Error", 
                f"Error al exportar: {str(e)}"
            )

    def cargar_inventario(self):
        try:
            conn = sqlite3.connect("db/store.db")
            cursor = conn.cursor()
            
            # Consulta con JOIN para obtener el nombre del usuario
            cursor.execute("""
                SELECT 
                    i.id, i.nombre, i.descripcion, i.cantidad,
                    i.precio_compra, i.precio_venta, i.categoria,
                    i.codigo_barras, i.fecha_creacion,
                    COALESCE(u.telefono, 'Sistema') as creador
                FROM inventario i
                LEFT JOIN usuarios u ON i.usuario_creacion = u.id
                ORDER BY i.id
            """)
            
            datos = cursor.fetchall()
            self.tabla_inventario.setRowCount(len(datos))
            
            for row, dato in enumerate(datos):
                for col, valor in enumerate(dato):
                    if col in [4, 5]:  # Valores monetarios
                        texto = f"${valor:.2f}"
                    elif col == 8:  # Fecha de creación
                        texto = valor.split()[0] if valor else ''  # Solo la fecha, sin hora
                    else:
                        texto = str(valor)
                    item = QTableWidgetItem(texto)
                    item.setTextAlignment(Qt.AlignCenter)
                    self.tabla_inventario.setItem(row, col, item)
            
            conn.close()
            
        except sqlite3.Error as e:
            QMessageBox.critical(self, "Error", f"Error al cargar el inventario: {str(e)}")

    def agregar_articulo(self):
        try:
            # Obtener valores de los campos
            nombre = self.input_widgets["Nombre:"].text().strip()
            descripcion = self.input_widgets["Descripción:"].text().strip()
            cantidad = self.input_widgets["Cantidad:"].text()
            valor_real = self.input_widgets["Valor Real:"].text()
            valor_venta = self.input_widgets["Valor Venta:"].text()
            categoria = self.input_widgets["Categoría:"].text().strip()

            # Validaciones
            if not nombre:
                raise ValueError("El nombre es obligatorio")
            if not cantidad.isdigit() or int(cantidad) < 0:
                raise ValueError("La cantidad debe ser un número positivo")
            try:
                valor_real = float(valor_real)
                valor_venta = float(valor_venta)
                if valor_real < 0 or valor_venta < 0:
                    raise ValueError()
            except:
                raise ValueError("Los valores deben ser números positivos")

            # Insertar en la base de datos
            conn = sqlite3.connect("db/store.db")
            cursor = conn.cursor()
            
            # Primero insertamos sin código de barras para obtener el ID
            cursor.execute("""
                INSERT INTO inventario (
                    nombre, descripcion, cantidad, 
                    precio_compra, precio_venta, 
                    categoria, usuario_creacion
                ) VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (
                nombre, descripcion, int(cantidad),
                valor_real, valor_venta,
                categoria, self.user_data['id']
            ))
            
            # Obtener el ID del artículo recién insertado
            articulo_id = cursor.lastrowid
            
            # Generar código de barras basado en el ID
            codigo_barras = f"{articulo_id:03d}"  # Formato: 001, 002, etc.
            
            # Actualizar el artículo con el código de barras
            cursor.execute("""
                UPDATE inventario 
                SET codigo_barras = ? 
                WHERE id = ?
            """, (codigo_barras, articulo_id))
            
            conn.commit()
            conn.close()
            
            # Limpiar campos y recargar tabla
            for input_widget in self.input_widgets.values():
                input_widget.clear()
            
            self.cargar_inventario()
            
            QMessageBox.information(
                self, 
                "Éxito", 
                f"Artículo agregado correctamente\nCódigo de barras asignado: {codigo_barras}"
            )
            
        except ValueError as e:
            QMessageBox.warning(self, "Error", str(e))
        except sqlite3.Error as e:
            QMessageBox.critical(self, "Error", f"Error en la base de datos: {str(e)}")

    def mostrar_editar_articulo(self):
        # Obtener la fila seleccionada
        current_row = self.tabla_inventario.currentRow()
        if current_row < 0:
            QMessageBox.warning(self, "Error", "Por favor seleccione un artículo para editar")
            return
        
        # Crear diálogo de edición
        dialog = EditarArticuloDialog(self, self.tabla_inventario, current_row)
        if dialog.exec_() == QDialog.Accepted:
            self.cargar_inventario()

    def editar_articulo(self):
        """Editar el artículo seleccionado en la base de datos."""
        try:
            id_articulo = self.input_id.text()
            nombre = self.input_nombre.text()
            descripcion = self.input_descripcion.text()
            cantidad = self.input_cantidad.text()
            valor_real = self.input_valor_real.text()
            valor_venta = self.input_valor_venta.text()

            if not (id_articulo.isdigit() and nombre and cantidad.isdigit() 
                    and valor_real.replace('.', '', 1).isdigit() 
                    and valor_venta.replace('.', '', 1).isdigit()):
                QMessageBox.warning(self, "Error", "Por favor, completa todos los campos correctamente.")
                return

            conn = sqlite3.connect("db/store.db")  # Cambiado a store.db
            cursor = conn.cursor()
            cursor.execute("""
            UPDATE inventario
            SET nombre = ?, descripcion = ?, cantidad = ?, precio_compra = ?, precio_venta = ?
            WHERE id = ?
            """, (nombre, descripcion, int(cantidad), float(valor_real), float(valor_venta), int(id_articulo)))
            
            conn.commit()
            conn.close()

            QMessageBox.information(self, "Éxito", "Artículo editado correctamente.")
            self.cargar_inventario()
            self.limpiar_campos()
            
        except sqlite3.Error as e:
            QMessageBox.critical(self, "Error", f"Error en la base de datos: {str(e)}")

    def eliminar_articulo(self):
        """Eliminar el artículo seleccionado de la base de datos."""
        try:
            id_articulo = self.input_id.text()

            if not id_articulo.isdigit():
                QMessageBox.warning(self, "Error", "Por favor, selecciona un artículo válido para eliminar.")
                return

            conn = sqlite3.connect("db/store.db")  # Cambiado a store.db
            cursor = conn.cursor()
            cursor.execute("DELETE FROM inventario WHERE id = ?", (int(id_articulo),))
            conn.commit()
            conn.close()

            QMessageBox.information(self, "Éxito", "Artículo eliminado correctamente.")
            self.cargar_inventario()
            self.limpiar_campos()
            
        except sqlite3.Error as e:
            QMessageBox.critical(self, "Error", f"Error en la base de datos: {str(e)}")

    def generar_codigo(self):
        """
        Generar código de barras para el artículo seleccionado por ID y enviarlo a la impresora
        (o guardarlo según esté implementado en imprimir_codigo_barras).
        """
        id_articulo = self.input_id.text()
        if not id_articulo.isdigit():
            QMessageBox.warning(self, "Error", "Por favor, ingresa un ID válido.")
            return

        imprimir_codigo_barras(int(id_articulo))
        QMessageBox.information(self, "Éxito", f"Código de barras generado e impreso para el ID {id_articulo}.")

    def limpiar_campos(self):
        """Deja en blanco los campos de entrada."""
        self.input_id.clear()
        self.input_nombre.clear()
        self.input_descripcion.clear()
        self.input_cantidad.clear()
        self.input_valor_real.clear()
        self.input_valor_venta.clear()

    def abrir_ventas(self):
        from frontend.sales_ui import SalesWindow
        self.ventana_ventas = SalesWindow(user_data=self.user_data, parent=self)
        self.ventana_ventas.show()


class EditarArticuloDialog(QDialog):
    def __init__(self, parent, tabla, row):
        super().__init__(parent)
        self.tabla = tabla
        self.row = row
        self.setup_ui()
        self.cargar_datos()

    def setup_ui(self):
        self.setWindowTitle("Editar Artículo")
        self.setMinimumWidth(400)
        layout = QVBoxLayout(self)
        
        # Formulario
        form_layout = QFormLayout()
        
        # ID y código de barras (no editables)
        self.id_input = QLineEdit()
        self.id_input.setReadOnly(True)
        form_layout.addRow("ID:", self.id_input)
        
        self.codigo_barras_input = QLineEdit()
        self.codigo_barras_input.setReadOnly(True)
        form_layout.addRow("Código Barras:", self.codigo_barras_input)
        
        # Campos editables
        self.nombre_input = QLineEdit()
        self.descripcion_input = QLineEdit()
        self.cantidad_input = QSpinBox()
        self.cantidad_input.setMaximum(99999)
        self.precio_compra_input = QDoubleSpinBox()
        self.precio_compra_input.setMaximum(999999.99)
        self.precio_compra_input.setDecimals(2)
        self.precio_venta_input = QDoubleSpinBox()
        self.precio_venta_input.setMaximum(999999.99)
        self.precio_venta_input.setDecimals(2)
        self.categoria_input = QLineEdit()
        
        form_layout.addRow("Nombre:", self.nombre_input)
        form_layout.addRow("Descripción:", self.descripcion_input)
        form_layout.addRow("Cantidad:", self.cantidad_input)
        form_layout.addRow("Precio Compra:", self.precio_compra_input)
        form_layout.addRow("Precio Venta:", self.precio_venta_input)
        form_layout.addRow("Categoría:", self.categoria_input)
        
        layout.addLayout(form_layout)
        
        # Botones
        button_box = QDialogButtonBox(
            QDialogButtonBox.Ok | QDialogButtonBox.Cancel
        )
        button_box.accepted.connect(self.guardar_cambios)
        button_box.rejected.connect(self.reject)
        layout.addWidget(button_box)

    def cargar_datos(self):
        # Cargar datos de la fila seleccionada
        self.id_input.setText(self.tabla.item(self.row, 0).text())
        self.nombre_input.setText(self.tabla.item(self.row, 1).text())
        self.descripcion_input.setText(self.tabla.item(self.row, 2).text())
        self.cantidad_input.setValue(int(self.tabla.item(self.row, 3).text()))
        self.precio_compra_input.setValue(float(self.tabla.item(self.row, 4).text().replace('$', '')))
        self.precio_venta_input.setValue(float(self.tabla.item(self.row, 5).text().replace('$', '')))
        self.categoria_input.setText(self.tabla.item(self.row, 6).text())
        self.codigo_barras_input.setText(self.tabla.item(self.row, 7).text())

    def guardar_cambios(self):
        try:
            # Validaciones
            if not self.nombre_input.text().strip():
                raise ValueError("El nombre no puede estar vacío")
            
            # Actualizar en la base de datos
            conn = sqlite3.connect("db/store.db")
            cursor = conn.cursor()
            
            cursor.execute("""
                UPDATE inventario 
                SET nombre = ?,
                    descripcion = ?,
                    cantidad = ?,
                    precio_compra = ?,
                    precio_venta = ?,
                    categoria = ?,
                    codigo_barras = ?
                WHERE id = ?
            """, (
                self.nombre_input.text().strip(),
                self.descripcion_input.text().strip(),
                self.cantidad_input.value(),
                self.precio_compra_input.value(),
                self.precio_venta_input.value(),
                self.categoria_input.text().strip(),
                self.codigo_barras_input.text().strip(),
                int(self.id_input.text())
            ))
            
            conn.commit()
            conn.close()
            
            QMessageBox.information(self, "Éxito", "Artículo actualizado correctamente")
            self.accept()
            
        except ValueError as e:
            QMessageBox.warning(self, "Error", str(e))
        except sqlite3.Error as e:
            QMessageBox.critical(self, "Error", f"Error en la base de datos: {str(e)}")


if __name__ == "__main__":
    app = QApplication([])
    window = InventoryApp()
    window.show()
    sys.exit(app.exec_())
